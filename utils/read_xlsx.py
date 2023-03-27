# Openpyxl
import openpyxl
from apps.bonos.models import Bono

def read_bonus():
    wb_obj = openpyxl.load_workbook('vitalicio.xlsx')
    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row
    max_col = sheet_obj.max_column
    
    for i in range(1, max_row+1):
        abonado = {
            'nombre': sheet_obj.cell(row = i, column = 2).value,
            'correo': '',
            'telefono': ''
        }
        
        tipo = sheet_obj.cell(row = i, column = 1).value
        
        estadio = {
            'seccion': sheet_obj.cell(row = i, column = 3).value,
            'fila': sheet_obj.cell(row = i, column = 4).value,
            'butaca': sheet_obj.cell(row = i, column = 5).value
        }        
        
        Bono.objects.create(
            tipo = tipo,
            abonado = abonado,
            ubicacion = {'estadio': estadio}
        )
        i = i+1