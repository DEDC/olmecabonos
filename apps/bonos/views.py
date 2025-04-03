# Python
import datetime
import json

from django.contrib.auth.mixins import LoginRequiredMixin
# Django
from django.views.generic import CreateView, ListView, RedirectView, TemplateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.db.models import Q, Count
from django.http import HttpResponseRedirect, JsonResponse
from django.core.cache import cache
from django.contrib import messages
from django.shortcuts import redirect
from django.views.decorators.csrf import csrf_exempt
from rest_framework.views import APIView

# app bonos
from apps.bonos.models import Bono, Partidos, Asistencias
from apps.bonos.forms import BonosForm, PartidosForm
# utils
from utils.bonos_pdf import generate_bonus, generate_qr
# openpyxl
from openpyxl import load_workbook
from django.http import HttpResponse
import pandas as pd
from io import BytesIO


class Registrar(LoginRequiredMixin, CreateView):
    model = Bono
    template_name = 'bonos/registro.html'
    form_class = BonosForm
    success_url = reverse_lazy('registro')
    
    def form_valid(self, form):
        abonado = {
            'name': self.request.POST.get('bn-name', 'ND').title(),
            'email': self.request.POST.get('bn-email', 'ND'),
            'phone': self.request.POST.get('bn-phone', 'ND')
        }
        
        bono = {
            'section': self.request.POST.get('bn-section', 'ND').upper(),
            'row': self.request.POST.get('bn-row', 'ND').upper(),
            'seat': self.request.POST.get('bn-seat', 'ND').upper()
        }
        
        pago = {
            'paytype': self.request.POST.get('bn-payment-type', 'ND'),
            'payto': self.request.POST.get('bn-payment-to', 'ND'),
            'payamount': self.request.POST.get('bn-payment-amount', 'ND')
        }
        bonos = Bono.objects.filter(ubicacion=bono, tipo=self.request.POST.get('tipo'))
        if bonos:
            messages.error(self.request, 'El asiento ya ha sido reservado con anterioridad')
            return HttpResponseRedirect(self.success_url)

        form.instance.abonado = abonado
        form.instance.ubicacion = bono
        form.instance.pago = pago
        self.object = form.save()
        messages.success(self.request, 'Bono registrado exitosamente')
        if self.request.POST.get('sv-dw', None) is not None:
            response_bn = generate_bonus([self.object])
            return response_bn
        elif self.request.POST.get('sv-qr', None) is not None:
            response_qr = generate_qr(self.object)
            return response_qr
        return HttpResponseRedirect(self.get_success_url())


class Listar(LoginRequiredMixin, ListView):
    model = Bono
    template_name = 'bonos/listado.html'
    paginate_by = 250
    
    def get_queryset(self):
        q = self.request.GET.get('q', '')
        dstart = self.request.GET.get('start', '')
        dend = self.request.GET.get('end', '')
        type_ = self.request.GET.getlist('type', [])
        lookup = (Q(folio__icontains = q) | Q(abonado__name__icontains = q))
        bonus = self.model._default_manager.filter(lookup)
        if len(type_) > 0:
            bonus = bonus.filter(tipo__in=type_)
        if dstart:
            try:
                if dend:
                    bonus = bonus.filter(fecha_reg__date__range=[datetime.datetime.fromisoformat(dstart).date(), datetime.datetime.fromisoformat(dend).date()])
                else:
                    bonus = bonus.filter(fecha_reg__date__range=[datetime.datetime.fromisoformat(dstart).date(), datetime.datetime.fromisoformat(dstart).date()])
            except:
                pass
        else:
            bonus = bonus.filter(fecha_reg__year=2024)
        self.queryset = bonus
        return bonus
    
    def get_context_data(self):
        type_ = ''
        for t in self.request.GET.getlist('type', []):
            type_+='&type={}'.format(t)
        context = {
            'q': self.request.GET.get('q', ''), 
            'type': type_, 
            'total': self.queryset.count(),
            'start': self.request.GET.get('start', ''),
            'end': self.request.GET.get('end', '')
        }
        return super().get_context_data(**context)


class Descargar(LoginRequiredMixin, RedirectView):
    url = reverse_lazy('listar')
    
    def get(self, request, *args, **kwargs):
        bonus = [request.GET.get('bonus')]
        try:
            bonos = Bono.objects.filter(folio__in=bonus)
            if bonos.exists():
                response = generate_bonus(bonos)
            return response
        except Exception as e:
            print(e)
        return super().get(request, *args, **kwargs)
    
    def post(self, request, *args, **kwargs):
        bonus = request.POST.getlist('bonus')
        try:
            bonos = Bono.objects.filter(folio__in=bonus)
            if bonos.exists():
                if 'dw-bn' in request.POST:
                    response = generate_bonus(bonos)
                if 'dw-qr' in request.POST:
                    response = generate_qr(bonos)
                return response
        except: pass
        return self.get(request, *args, **kwargs)


class CargarExcel(LoginRequiredMixin, TemplateView):
    template_name = 'bonos/carga_excel.html'
    
    def get(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        context['bonus'] = cache.get('bonus_cache')
        return self.render_to_response(context)
    
    def post(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        if request.POST.get('upload') is not None:
            file_ = request.FILES.get('file')
            wb = load_workbook(file_)
            ws = wb.active
            objs = []
            for row in ws.iter_rows(min_row=2):
                abonado = {'name': row[0].value.upper(), 'email': '', 'phone': ''}
                bono = {'section': str(row[1].value).upper(), 'row': str(row[2].value).upper(), 'seat': str(row[3].value).upper()}
                # if row[5].value is not None:
                #     bono.update({'extra': str(row[5].value)})
                objs.append(Bono(abonado=abonado, ubicacion=bono, tipo=row[4].value.lower()))
            cache.set('bonus_cache', objs)
        if request.POST.get('save') is not None:
            bonus_cache = cache.get('bonus_cache')
            if isinstance(bonus_cache, list):
                for obj in bonus_cache:
                    obj.save()
                # bonus_saved = Bono.objects.bulk_create(cache.get('bonus_cache'))
                messages.success(self.request, 'Se registraron {} bono(s) exitosamente'.format(len(bonus_cache)))
                cache.set('bonus_cache', None)
            else:
                if bonus_cache:
                    messages.error(self.request, 'No se registraron {} bono(s) exitosamente'.format(len(bonus_cache)))
                else:
                    messages.error(self.request, 'No hay bono(s) que cargar')
        context['bonus'] = cache.get('bonus_cache')
        return self.render_to_response(context)


class Editar(LoginRequiredMixin, UpdateView):
    template_name = 'bonos/editar.html'
    model = Bono
    form_class = BonosForm
    slug_field = 'uuid'
    slug_url_kwarg = 'uuid'
    
    def form_valid(self, form):
        abonado = {
            'name': self.request.POST.get('bn-name', 'ND').title(),
            'email': self.request.POST.get('bn-email', 'ND'),
            'phone': self.request.POST.get('bn-phone', 'ND')
        }
        bono = {
            'section': self.request.POST.get('bn-section', 'ND').upper(),
            'row': self.request.POST.get('bn-row', 'ND').upper(),
            'seat': self.request.POST.get('bn-seat', 'ND').upper()
        }
        
        pago = {
            'paytype': self.request.POST.get('bn-payment-type', 'ND'),
            'payto': self.request.POST.get('bn-payment-to', 'ND'),
            'payamount': self.request.POST.get('bn-payment-amount', 'ND')
        }
        
        if abonado != self.object.abonado or bono != self.object.ubicacion or pago != self.object.pago:
            form.instance.abonado = abonado
            form.instance.ubicacion = bono
            form.instance.pago = pago
            self.object = form.save()
            messages.success(self.request, 'Bono editado exitosamente')
            if self.request.POST.get('sv-dw', None) is not None:
                response_bn = generate_bonus([self.object])
                return response_bn
            elif self.request.POST.get('sv-qr', None) is not None:
                response_qr = generate_qr(self.object)
                return response_qr
        else:
            messages.warning(self.request, 'No se detectó nigún cambio en la información')
        return HttpResponseRedirect(self.get_success_url())

    def get_success_url(self):
        return reverse_lazy('editar', kwargs={'uuid': self.object.uuid})


class Eliminar(LoginRequiredMixin, DeleteView):
    template_name = 'bonos/eliminar.html'
    model = Bono
    slug_field = 'uuid'
    slug_url_kwarg = 'uuid'
    success_url = reverse_lazy('listar')
    
    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Bono eliminado exitosamente')
        return super().delete(request, *args, **kwargs)


class Lector(LoginRequiredMixin, TemplateView):
    template_name = 'bonos/lector.html'
    
    def get(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        try:
            partido = Partidos.objects.get(uuid=kwargs.get('uuid'))
        except Partidos.DoesNotExist:
            return redirect('juegos')
        context['partido'] = partido
        return self.render_to_response(context)


class Juegos(LoginRequiredMixin, CreateView):
    template_name = 'bonos/juegos.html'
    model = Partidos
    form_class = PartidosForm
    success_url = reverse_lazy('juegos')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["partidos"] = Partidos.objects.annotate(count=Count('asistencias_partido', filter=Q(asistencias_partido__bono__tipo__in=['abonado', 'palco']))).order_by('-fecha_reg')
        return context


@csrf_exempt
def check_bonus(request):
    # request should be ajax and method should be GET.
    if request.is_ajax and request.method == "POST":
        data = json.loads(request.body)
        try:
            bonus = data.get('bonus', '')
            partido_txt = data.get('partido', '')
            bono = Bono.objects.get(folio__exact=bonus)
            partido = Partidos.objects.get(uuid=partido_txt)
            asis, created = Asistencias.objects.get_or_create(bono=bono, partido=partido)
            response = {
                'ubication': bono.ubicacion,
                'person': bono.abonado,
                'created': created
            }
            return JsonResponse(response, status = 200)
        except Exception as e:
            print(e)
    return JsonResponse({}, status = 400)


class GenerateExcelApiView(LoginRequiredMixin, APIView):
    def get(self, request, *args, **kwargs):
        # Obtener todos los datos del modelo Person
        bonos_list = Bono.objects.all()

        if self.request.GET.get('start'):
            bonos_list = bonos_list.filter(fecha_reg__gte=self.request.GET.get('start'))

        if self.request.GET.get('stop'):
            bonos_list = bonos_list.filter(fecha_reg__lte=self.request.GET.get('stop'))

        if self.request.GET.get('type'):
            bonos_list = bonos_list.filter(tipo__in=self.request.GET.get('type').split(','))

        data = []

        for bono in bonos_list:
            data.append({
                "folio": bono.folio,
                "tipo": bono.tipo,
                "nombre": bono.abonado.get("name"),
                "telefono": bono.abonado.get("phone"),
                "correo": bono.abonado.get("email"),
                "seccion": bono.ubicacion.get("section"),
                "fila": bono.ubicacion.get("row"),
                "asiento": bono.ubicacion.get("seat"),
                "tipo_pago": bono.pago.get("paytype"),
                "pago": bono.pago.get("payamount"),
            })

        # Crear un DataFrame de pandas con los datos
        df = pd.DataFrame(data)

        # Crear un buffer en memoria para guardar el archivo Excel
        excel_buffer = BytesIO()

        # Guardar el DataFrame en el buffer como un archivo Excel
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Bonos')

        # Mover el cursor del buffer al inicio
        excel_buffer.seek(0)

        # Crear la respuesta HTTP con el archivo Excel adjunto
        response = HttpResponse(excel_buffer,
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="bonos.xlsx"'

        return response